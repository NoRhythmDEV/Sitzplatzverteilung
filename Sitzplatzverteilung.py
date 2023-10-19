import tkinter as tk
from tkinter import WORD, filedialog, messagebox
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
import os
import random
import time


# Funktion zum Importieren von Daten aus einer Excel-Datei
def import_from_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            data = [str(sheet.cell(row=i, column=1).value).strip() for i in range(1, sheet.max_row + 1)]
            # Daten mit Komma trennen
            data = ','.join(data)
            # Daten im Textfeld anzeigen
            entry_namen.delete("1.0", tk.END)
            entry_namen.insert(tk.END, data)
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Importieren der Daten aus der Excel-Datei:\n{str(e)}")


#Export als PDF
def save_as_pdf():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Get the current timestamp in the format HHMMSS
    timestr = time.strftime("%Hh-%Mm-%Ss")

    # Original filename
    original_filename = "Sitzordnung.pdf"

    # Extract the filename and extension
    filename, extension = os.path.splitext(original_filename)

    # Append the timestamp to the filename
    new_filename = f"{filename}_{timestr}{extension}"

    file_path = os.path.join(desktop_path, new_filename)
    text_content = text_output.get("1.0", tk.END).strip()

    pdf = canvas.Canvas(file_path, pagesize=letter)
    pdf.setFont("Helvetica", 10)
    
    # Split the text into lines
    lines = text_content.split('\n')

    # Define page parameters
    max_lines = 45  # Maximum lines per page
    current_line = 0
    page_number = 1

    # Loop through the lines and add them to the PDF
    for line in lines:
        if current_line >= max_lines:
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            current_line = 0
            page_number += 1
        y_position = 730 - (current_line * 15)
        pdf.drawString(50, y_position, line)
        current_line += 1

    pdf.save()

    messagebox.showinfo("PDF gespeichert", f"Die Datei wurde als '{new_filename}' auf dem Desktop gespeichert (Seite {page_number}).")

# Funktion zum Generieren der Sitzordnungen
def generate_sitzordnungen():
    namen_liste = entry_namen.get("1.0", tk.END).strip()
    personen_pro_tisch = int(entry_personen_pro_tisch.get())

    # Namen in eine Liste aufteilen und bereinigen
    namen = [name.strip() for name in namen_liste.split(',')]

    # Zufällige Sitzordnung generieren
    random.shuffle(namen)

    # Anzahl der Tische berechnen
    anzahl_tische = len(namen) // personen_pro_tisch
    restliche_personen = len(namen) % personen_pro_tisch

    # Output-Text für beide Sitzordnungen
    sitzordnungen_text = ""

    for ordnung_nummer in range(2):
        sitzordnungen_text += f"Sitzordnung {ordnung_nummer + 1}:\n\n"
        tisch_counter = 1
        personen_counter = 0

        for i in range(anzahl_tische):
            tisch_personen = []
            for j in range(personen_pro_tisch):
                tisch_personen.append(namen[personen_counter])
                personen_counter += 1

            sitzordnungen_text += f"[Tisch {tisch_counter}]\n\n [{' | '.join(tisch_personen)}]\n\n"
            tisch_counter += 1
#→
        # Restliche Personen hinzufügen
        if restliche_personen > 0:
            tisch_personen = []
            for k in range(restliche_personen):
                tisch_personen.append(namen[personen_counter])
                personen_counter += 1

            sitzordnungen_text += f"[Tisch {tisch_counter}]\n\n [{' | '.join(tisch_personen)}]\n\n"

    # Output im Textfeld anzeigen
    text_output.delete("1.0", tk.END)
    text_output.insert(tk.END, sitzordnungen_text)

# Funktion zum Schließen des Fensters
def close_window():
    root.destroy()

# GUI erstellen
root = tk.Tk()
root.title("Sitzordnungen Generator")
root.configure(background="#424242")

# Mindestgröße setzen
root.minsize(1600, 900)
root.attributes("-fullscreen", True)

# Eingabefeld für Namen
label_namen = tk.Label(root, font=("Arial", 16), text="Namen (getrennt durch Komma):")
label_namen.grid(row=0, column=0, padx=(15, 10), pady=(15, 10), sticky="w")

entry_namen = tk.Text(root, font=("Arial", 12), width=25, height=40, wrap=WORD)
entry_namen.grid(row=1, column=0, padx=(15, 10), pady=(0, 10), sticky="nsew")

# Ausgabefeld für Sitzordnungen
label_ausgabe = tk.Label(root, font=("Arial", 16), text="Ausgabe:")
label_ausgabe.grid(row=0, column=1, padx=(15, 10), pady=(15, 10), sticky="w",)

text_output = tk.Text(root, font=("Arial", 12), width=120, height=40)
text_output.grid(row=1, column=1, padx=(15, 10), pady=(0, 10), sticky="nsew")

# Eingabefeld für Personen pro Tisch
label_personen_pro_tisch = tk.Label(root, font=("Arial", 16), text="Personen pro Tisch (Ganzzahl):")
label_personen_pro_tisch.grid(row=2, column=0, padx=(15, 10), pady=(15, 10), sticky="w")

entry_personen_pro_tisch = tk.Entry(root, font=("Arial", 12), width=25, justify="center")
entry_personen_pro_tisch.grid(row=3, column=0, padx=(15, 10), pady=(0, 10), sticky="nswe")

# Import Excel-Sheet
importexcel_button = tk.Button(root, font=("Arial", 16), width=50, text="1. Infos aus Excel-Datei importieren", command=import_from_excel)
importexcel_button.grid(row=4, column=0, padx=(15, 10), pady=(10, 15))

# Button zum Generieren der Sitzordnungen
generate_button = tk.Button(root, font=("Arial", 16), text="2. Sitzordnungen generieren", command=generate_sitzordnungen, width=50)
generate_button.grid(row=5, column=0, padx=(15, 10), pady=(10, 15))

# Als PDF Speichern Button
saveaspdf_button = tk.Button(root, font=("Arial", 16), width=50, text="3. Als PDF Speichern", command=save_as_pdf)
saveaspdf_button.grid(row=4, column=1, padx=(15, 10), pady=(10, 15))

# Exit-Button
exit_button = tk.Button(root, font=("Arial", 16), width=50, text="4. Exit", command=close_window)
exit_button.grid(row=5, column=1, padx=(15, 10), pady=(10, 15))

# GUI starten
root.mainloop()